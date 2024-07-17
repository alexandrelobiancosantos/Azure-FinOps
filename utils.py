import json
import logging
#import statistics
import subprocess
import time
from datetime import datetime, timedelta

import pandas as pd
import requests
from tabulate import tabulate
import openpyxl
from openpyxl.chart import PieChart, Reference
#from openpyxl.drawing.image import Image
from openpyxl.chart.label import DataLabelList

# Contador global de requisições e tempo da última requisição
request_count = 0
last_request_time = None

def increment_request_count():
    global request_count, last_request_time
    request_count += 1
    current_time = time.time()
    if last_request_time is not None:
        interval = current_time - last_request_time
        logging.info(f"Number of requests made: {request_count} | Interval since last request: {interval:.2f} seconds")
    else:
        logging.info(f"Number of requests made: {request_count} | This is the first request.")
    last_request_time = current_time

def setup_logging():
    """Set up basic logging configuration."""
    logging.basicConfig(level=logging.INFO)
 
def handle_errors(exception, message):
    """Handle errors by logging the message and exception, then exiting the program."""
    logging.error(f"{message}: {exception}")
    exit(1)
 
def find_common_prefix(strings):
    """Find the longest common prefix among a list of strings."""
    if not strings:
        return ""
    shortest_str = min(strings, key=len)
    for i, char in enumerate(shortest_str):
        for other in strings:
            if other[i] != char:
                return shortest_str[:i]
    return shortest_str
 
def get_subscription_ids(subscription_prefix):
    """
    Retrieve subscription IDs that start with the given prefix.
    Returns:
        list of tuples: List of subscription names and IDs.
    """
    try:
        result = subprocess.run(
            ["az", "account", "list"],
            capture_output=True,
            text=True,
            check=True
        )
        increment_request_count()  # Incrementa o contador de requisições
        subscriptions = json.loads(result.stdout)
        subscription_ids = [
            (subscription['name'], subscription['id'])
            for subscription in subscriptions
            if subscription['name'].startswith(subscription_prefix)
        ]
        if not subscription_ids:
            logging.error(f"No subscriptions found with prefix '{subscription_prefix}'.")
            exit(1)
        return subscription_ids
    except subprocess.CalledProcessError as e:
        handle_errors(e, "Command error")
    except json.JSONDecodeError as e:
        handle_errors(e, "JSON decode error")
    except Exception as e:
        handle_errors(e, "Unexpected error")
 
def get_access_token():
    """
    Retrieve an access token for the Azure management API.
 
    Returns:
        str: The access token.
    """
    try:
        result = subprocess.run(
            ["az", "account", "get-access-token", "--resource=https://management.azure.com/"],
            capture_output=True,
            text=True,
            check=True
        )
        increment_request_count()  # Incrementa o contador de requisições
        token_info = json.loads(result.stdout)
        return token_info['accessToken']
    except subprocess.CalledProcessError as e:
        handle_errors(e, "Command error")
    except json.JSONDecodeError as e:
        handle_errors(e, "JSON decode error")
    except Exception as e:
        handle_errors(e, "Unexpected error")
 
def get_analysis_timeframe(start_date_str=None, period=31):
    """
    Get the analysis timeframe retroactive to seven days from the given date or yesterday if no date is given.
    Args:
        start_date_str (str, optional): The start date in 'YYYY-MM-DD' format. Defaults to None.
 
    Returns:
        tuple: Start date, end date, and timeframe dictionary.
    """
    if start_date_str:
        end_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    else:
        end_date = datetime.utcnow() - timedelta(days=1)
    start_date = end_date - timedelta(days=period)
    timeframe = {
        "from": start_date.strftime('%Y-%m-%d'),
        "to": end_date.strftime('%Y-%m-%d')
    }
    return start_date, end_date, timeframe
 
def build_cost_management_request(subscription_id, grouping_type, grouping_name, access_token):
    cost_management_url = f'https://management.azure.com/subscriptions/{subscription_id}/providers/Microsoft.CostManagement/query?api-version=2021-10-01'
    start_date, end_date, timeframe = get_analysis_timeframe()
    payload = {
        "type": "ActualCost",
        "timeframe": "Custom",
        "timePeriod": timeframe,
        "dataset": {
            "granularity": "Daily",
            "aggregation": {
                "totalCost": {
                    "name": "Cost",
                    "function": "Sum"
                }
            },
            "grouping": []
        }
    }
    if grouping_type.lower() != 'subscription':
        payload["dataset"]["grouping"].append({
            "type": grouping_type,
            "name": grouping_name
        })
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }
    return cost_management_url, payload, headers

def check_alert(cost_yesterday, average_cost):
    """
    Check if the cost for yesterday exceeds the average cost.
    Returns:
        str: "Yes" if cost_yesterday exceeds average_cost, otherwise "No".
    """
    return "Yes" if cost_yesterday > (average_cost + 0.01) else "No"
 
def process_costs(costs_by_group, grouping_key, start_date, end_date, analysis_date_str):
    """
    Process costs by group and calculate average costs, alerts, and additional metrics.

    Args:
        costs_by_group (dict): The costs grouped by a specific key.
        grouping_key (str): The key to group costs by.
        start_date (datetime): The start date for the analysis period.
        end_date (datetime): The end date for the analysis period.
        analysis_date_str (str): The string representation of the analysis date.

    Returns:
        list: List of results with average costs, alerts, and additional metrics.
    """
    results = []
    analysis_date = datetime.strptime(analysis_date_str, '%Y%m%d')
    is_analysis_date_weekend = analysis_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

    for group_value, costs in costs_by_group.items():
        weekday_costs = []
        weekend_costs = []

        for single_date in (start_date + timedelta(n) for n in range((end_date - start_date).days + 1)):
            date_int = int(single_date.strftime('%Y%m%d'))
            day_cost = next((cost for date, cost in costs if date == date_int), 0)
            if single_date.weekday() >= 5:  # Saturday or Sunday
                weekend_costs.append(day_cost)
            else:  # Weekday
                weekday_costs.append(day_cost)

        if is_analysis_date_weekend:
            total_days = len(weekend_costs)
            total_cost = sum(weekend_costs)
        else:
            total_days = len(weekday_costs)
            total_cost = sum(weekday_costs)

        average_cost = total_cost / total_days if total_days > 0 else 0
        cost_on_analysis_date = next((cost for date, cost in costs if date == int(analysis_date_str)), 0)
        alert = check_alert(cost_on_analysis_date, average_cost)
        percent_variation = ((cost_on_analysis_date - average_cost) / average_cost) * 100 if average_cost != 0 else 0
        cost_difference = cost_on_analysis_date - average_cost

        results.append({
            grouping_key: group_value,
            "Average Cost": average_cost,
            "Analysis Date Cost": cost_on_analysis_date,
            "Alert": alert,
            "Percent Variation": percent_variation,
            "Cost Difference": cost_difference,
            "Period of Average Calculation": f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
            "Number of Days": total_days,
            "Analysis Date": end_date.strftime('%Y-%m-%d')
        })

    return results


def request_and_process(url, headers, payload, subscription_name):
    """
    Send request to the Azure Cost Management API and process the response.
 
    Args:
        url (str): The API URL.
        headers (dict): The request headers.
        payload (dict): The request payload.
        subscription_name (str): The name of the subscription.
        Returns:
        dict or None: The response data or None if no cost found.
    """
    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        increment_request_count()  # Incrementa o contador de requisições
        time.sleep(1)  # Pausa de 1 segundos entre as requisições
    except requests.exceptions.RequestException as e:
        handle_errors(e, f"Failed to retrieve cost data for subscription '{subscription_name}'")
 
    try:
        data = response.json()
        logging.debug(f"Received data: {json.dumps(data, indent=2)}")
    except json.JSONDecodeError as e:
        handle_errors(e, "JSON decode error")
 
    if 'properties' not in data or 'rows' not in data['properties']:
        logging.info("No Cost Found in the response data.")
        return None, 0
 
    return data
 
def analyze_costs(subscription_name, subscription_id, grouping_dimension, access_token, start_date_str=None, period=31):
    """
    Analyze costs for a subscription grouped by a specific dimension.
    Args:
        subscription_name (str): Name of the subscription.
        subscription_id (str): ID of the subscription.
        grouping_dimension (str): The dimension to group costs by.
        access_token (str): Azure access token.
        start_date_str (str, optional): The start date for the analysis period. Defaults to None.
        period (int, optional): Number of days for the analysis period. Defaults to 31.

    Returns:
        tuple: Analysis result as a table, total cost on the analysis date, and the dataframe.
    """
    cost_management_url, payload, headers = build_cost_management_request(subscription_id, 'Dimension', grouping_dimension, access_token)

    start_date, end_date, _ = get_analysis_timeframe(start_date_str, period)

    logging.debug(f"Sending request to Cost Management API for subscription {subscription_id} with payload: {json.dumps(payload, indent=2)}")

    data = request_and_process(cost_management_url, headers, payload, subscription_name)

    if data is None:
        return "No Cost Found", 0, None

    costs_by_group = {}
    total_cost_analysis_date = 0
    analysis_date_str = end_date.strftime('%Y%m%d')

    for result in data['properties']['rows']:
        cost = float(result[0])
        date = result[1]
        group = result[2]

        if group not in costs_by_group:
            costs_by_group[group] = []
        costs_by_group[group].append((date, cost))

        if date == int(analysis_date_str):
            total_cost_analysis_date += cost

    results = process_costs(costs_by_group, grouping_dimension, start_date, end_date, analysis_date_str)

    df = pd.DataFrame(results)

    if df.empty:
        logging.info("No data to display.")
        return "No Cost Found", total_cost_analysis_date, None

    table = tabulate(df, headers='keys', tablefmt='plain', floatfmt='.3f')
    return table, total_cost_analysis_date, df

def analyze_costs_by_tag(subscription_name, subscription_id, tag_key, access_token, start_date_str=None, period=31):
    """
    Analyze costs for a subscription grouped by a specific tag key.
    Args:
        subscription_name (str): Name of the subscription.
        subscription_id (str): ID of the subscription.
        tag_key (str): The tag key to group costs by.
        access_token (str): Azure access token.
        start_date_str (str, optional): The start date for the analysis period. Defaults to None.
        period (int, optional): Number of days for the analysis period. Defaults to 31.

    Returns:
        tuple: Analysis result as a table, total cost on the analysis date, and the dataframe.
    """
    cost_management_url, payload, headers = build_cost_management_request(subscription_id, 'TagKey', tag_key, access_token)

    start_date, end_date, _ = get_analysis_timeframe(start_date_str, period)

    logging.debug(f"Sending request to Cost Management API for subscription {subscription_id} with payload: {json.dumps(payload, indent=2)}")

    data = request_and_process(cost_management_url, headers, payload, subscription_name)

    if data is None:
        return "No Cost Found", 0, None

    costs_by_tag = {}
    total_cost_analysis_date = 0
    analysis_date_str = end_date.strftime('%Y%m%d')

    for result in data['properties']['rows']:
        cost = float(result[0])
        date = result[1]
        tag_value = result[3]

        if tag_value:
            if tag_value not in costs_by_tag:
                costs_by_tag[tag_value] = []
            costs_by_tag[tag_value].append((date, cost))

            if date == int(analysis_date_str):
                total_cost_analysis_date += cost

    results = process_costs(costs_by_tag, tag_key, start_date, end_date, analysis_date_str)

    df = pd.DataFrame(results)

    if df.empty:
        logging.info("No data to display.")
        return "No Cost Found", total_cost_analysis_date, None

    table = tabulate(df, headers='keys', tablefmt='plain', floatfmt='.3f')
    return table, total_cost_analysis_date, df

def analyze_costs_by_subs(subscription_name, subscription_id, access_token, start_date_str=None, period=31):
    cost_management_url, payload, headers = build_cost_management_request(subscription_id, 'Subscription', subscription_name, access_token)
    start_date, end_date, _ = get_analysis_timeframe(start_date_str, period)
    logging.debug(f"Sending request to Cost Management API for subscription {subscription_id} with payload: {json.dumps(payload, indent=2)}")
    data = request_and_process(cost_management_url, headers, payload, subscription_name)
    if data is None:
        return {
            "Subscription": subscription_name,
            "Average Cost": 0,
            "Analysis Date Cost": 0,
            "Alert": "No",
            "Percent Variation": 0,
            "Cost Difference": 0,
            "Period of Average Calculation": f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
            "Number of Days": 0,
            "Analysis Date": end_date.strftime('%Y-%m-%d')
        }
    costs = []
    analysis_date_str = end_date.strftime('%Y%m%d')
    for result in data['properties']['rows']:
        cost = float(result[0])
        date = result[1]
        costs.append((date, cost))
    total_cost = sum(cost for date, cost in costs)
    total_days = len(costs)
    average_cost = total_cost / total_days if total_days > 0 else 0
    cost_on_analysis_date = next((cost for date, cost in costs if date == int(analysis_date_str)), 0)
    alert = check_alert(cost_on_analysis_date, average_cost)
    percent_variation = ((cost_on_analysis_date - average_cost) / average_cost) * 100 if average_cost != 0 else 0
    cost_difference = cost_on_analysis_date - average_cost
    return {
        "Subscription": subscription_name,
        "Average Cost": average_cost,
        "Analysis Date Cost": cost_on_analysis_date,
        "Alert": alert,
        "Percent Variation": percent_variation,
        "Cost Difference": cost_difference,
        "Period of Average Calculation": f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
        "Number of Days": total_days,
        "Analysis Date": end_date.strftime('%Y-%m-%d')
    }

def analyze_subscription(subscription_name, subscription_id, analysis_type, grouping_key, access_token, alert_mode=False, start_date_str=None, period=31):
    logging.info(f"\nAnalyzing subscription: {subscription_name} with ID: {subscription_id}")
    
    if analysis_type.lower() == 'tag':
        result, cost_analysis_date, df = analyze_costs_by_tag(subscription_name, subscription_id, grouping_key, access_token, start_date_str, period)
    elif analysis_type.lower() == 'group':
        result, cost_analysis_date, df = analyze_costs(subscription_name, subscription_id, grouping_key, access_token, start_date_str, period)
    else:  # For subscription analysis
        result = analyze_costs_by_subs(subscription_name, subscription_id, access_token, start_date_str, period)
        df = pd.DataFrame([result])
    
    if df is not None:
        if alert_mode:
            alert_df = df[df['Alert'] == 'Yes']
            if not alert_df.empty:
                logging.info(f"Alerts found for {subscription_name}.")
                result = tabulate(alert_df, headers='keys', tablefmt='plain', floatfmt='.3f')
                return subscription_name, alert_df, result
            else:
                logging.info(f"No alerts found for {subscription_name}.")
                return subscription_name, None, "No alerts found"
        else:
            result = tabulate(df, headers='keys', tablefmt='plain', floatfmt='.3f')
            return subscription_name, df, result
    else:
        return subscription_name, None, "No data found"

def save_execution_result(status, subscription_results, common_prefix, grouping_key):
    """Save the analysis result to an Excel file with each subscription in a separate sheet and generate pie charts."""
    timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
    filename = f"{common_prefix}_{grouping_key}_{timestamp}.xlsx"
    try:
        # Save the data to Excel using pandas
        with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
            for subscription_name, df in subscription_results.items():
                df.to_excel(writer, sheet_name=subscription_name, index=False, float_format="%.2f")

        # Load the saved Excel file using openpyxl
        wb = openpyxl.load_workbook(filename)
        
        for subscription_name in subscription_results.keys():
            ws = wb[subscription_name]
            # Create the first pie chart for Average Cost
            pie1 = PieChart()
            labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            pie1.add_data(data, titles_from_data=True)
            pie1.set_categories(labels)
            pie1.title = "Average Cost Distribution"
            pie1.dataLabels = DataLabelList()
            pie1.dataLabels.showPercent = False
            pie1.dataLabels.showSerName = False
            pie1.dataLabels.showCatName = False
            pie1.dataLabels.showVal = True

            # Place the first pie chart on the worksheet
            ws.add_chart(pie1, "A10")

            # Create the second pie chart for Analysis Date Cost
            pie2 = PieChart()
            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            pie2.add_data(data, titles_from_data=True)
            pie2.set_categories(labels)
            analysis_date = ws.cell(row=2, column=9).value  # Assumes Analysis Date is in the first row, column 10
            pie2.title = f"Cost Distribution on {analysis_date}"
            pie2.dataLabels = DataLabelList()
            pie2.dataLabels.showPercent = False
            pie2.dataLabels.showSerName = False
            pie2.dataLabels.showCatName = False
            pie2.dataLabels.showVal = True

            # Place the second pie chart on the worksheet
            ws.add_chart(pie2, "J10")

        # Save the modified Excel file
        wb.save(filename)
        logging.info(f"Results saved to {filename}")
    except Exception as e:
        logging.error(f"Failed to save results: {e}")
